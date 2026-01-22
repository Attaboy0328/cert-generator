import streamlit as st
import pandas as pd
from docxtpl import DocxTemplate
import io
import os
from docx import Document
from docxcompose.composer import Composer
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components

# --- 1. 页面基础配置 ---
st.set_page_config(page_title="证书智能制作工具", layout="centered")

# --- 2. 深度界面定制与主题切换逻辑 ---
def apply_theme_and_styles(is_dark):
    # 根据主题选择颜色
    bg_color = "#0E1117" if is_dark else "#FFFFFF"
    text_color = "#FAFAFA" if is_dark else "#31333F"
    footer_text_color = "#888" if is_dark else "#666"
    icon_opacity = "0.5" if is_dark else "0.7"
    border_color = "#30363D" if is_dark else "#E6E9EF"

    # A. 注入 CSS：修复边框、白条、并彻底居中标题
    st.markdown(f"""
        <style>
        /* 1. 基础容器透明化：解决“带边框”和“白条”问题 */
        .stApp, header[data-testid="stHeader"], .st-emotion-cache-6qob1r {{
            background-color: {bg_color} !important;
            color: {text_color} !important;
        }}
        
        /* 移除某些版本可能出现的容器边框或阴影 */
        div[data-testid="stVerticalBlock"] > div {{
            background-color: transparent !important;
            border: none !important;
            box-shadow: none !important;
        }}

        /* 2. 隐藏官方按钮 */
        div[data-testid="stStatusWidget"] {{ display: none !important; }}
        footer {{ visibility: hidden !important; }}
        
        /* 3. 标题强制居中与响应式 */
        .stApp h1 {{
            text-align: center !important;
            display: flex; justify-content: center; align-items: center;
            gap: 10px; width: 100%; margin-top: 0px;
        }}
        @media (max-width: 640px) {{
            .stApp h1 {{ font-size: 1.6rem !important; }}
            .stApp .block-container {{ padding: 1rem !important; }}
        }}
        .main .block-container {{ padding-bottom: 120px; }}
        
        /* 4. 优化输入框和表格在主题下的表现 */
        input, select, textarea, .stDataEditor {{
            border-color: {border_color} !important;
        }}
        </style>
    """, unsafe_allow_html=True)

    # B. 注入 JS：Logo跳转
    components.html("""
        <script>
        const targetUrl = "https://share.streamlit.io/user/attaboy0328";
        const logoSvg = `<svg xmlns="http://www.w3.org/2000/svg" width="22" height="22" viewBox="0 0 24 24" fill="#FF4B4B" style="margin-right:15px; cursor:pointer;"><path d="M12 2L2 19.72L12 22L22 19.72L12 2ZM12 16.5L6.5 15.5L12 6L17.5 15.5L12 16.5Z"/></svg>`;
        function injectLogo() {
            const header = window.parent.document.querySelector('header[data-testid="stHeader"]');
            const container = header ? header.querySelector('div:nth-child(2)') : null;
            if (container && !window.parent.document.getElementById('custom-streamlit-logo')) {
                const link = window.parent.document.createElement('a');
                link.id = 'custom-streamlit-logo';
                link.href = targetUrl; link.target = "_blank";
                link.innerHTML = logoSvg; link.style.display = "flex"; link.style.alignItems = "center";
                container.prepend(link);
            }
        }
        setInterval(injectLogo, 500);
        </script>
    """, height=0)

    # 返回底部HTML
    return f"""
    <div style="text-align:center;margin-top:50px;padding-bottom:20px;width:100%;">
        <div style="display:flex;justify-content:center;align-items:center;gap:20px;margin-bottom:15px;flex-wrap:wrap;">
            <img src="https://cdn.jsdelivr.net/gh/devicons/devicon/icons/github/github-original.svg" width="22" style="opacity:{icon_opacity};filter:grayscale(1);">
            <img src="https://www.vectorlogo.zone/logos/cloudflare/cloudflare-ar21.svg" width="55" style="opacity:{icon_opacity};">
            <img src="https://www.vectorlogo.zone/logos/vercel/vercel-ar21.svg" width="55" style="opacity:{icon_opacity};">
            <img src="https://cdn.jsdelivr.net/gh/devicons/devicon/icons/vuejs/vuejs-original.svg" width="22" style="opacity:{icon_opacity};">
            <img src="https://www.vectorlogo.zone/logos/tailwindcss/tailwindcss-icon.svg" width="22" style="opacity:{icon_opacity};">
        </div>
        <div style="font-size:12px;color:{footer_text_color};line-height:1.6;font-family:sans-serif;letter-spacing:0.5px;">
            <p style="margin:0;">© 2026 Jiachen Tu. All rights reserved.</p>
        </div>
    </div>
    """

# --- 3. 页面渲染 ---

# 放置开关
col_left, col_right = st.columns([8, 2])
with col_right:
    is_dark_mode = st.toggle("🌙 夜间模式", value=False)

# 应用样式并获取底部内容
footer_html = apply_theme_and_styles(is_dark_mode)

st.markdown("<h1>🎓 内审员证书智能制作工具</h1>", unsafe_allow_html=True)

# 业务内容
st.markdown("### 第一步：选择录入模式")
mode = st.radio("选择方式：", ["Excel 文件上传", "网页表格填写 (支持粘贴)"], index=0, horizontal=True)

DEFAULT_TEMPLATE = "内审员证书.docx"
data_to_process = []

st.markdown("---")
st.markdown("### 第二步：填写或上传信息")

if mode == "网页表格填写 (支持粘贴)":
    st.info("💡 提示：点击左上角第一个单元格并按下 Ctrl+V 即可粘贴 Excel 数据。")
    init_df = pd.DataFrame({
        "序号": [i for i in range(1, 101)],
        "证书编号": [None] * 100, "姓名": [None] * 100, "身份证号": [None] * 100, "培训日期": [None] * 100, "标准号": [None] * 100,
    })
    edited_df = st.data_editor(
        init_df, num_rows="fixed", use_container_width=True, hide_index=True, height=380,
        column_config={
            "序号": st.column_config.NumberColumn("序号", width=40, disabled=True),
            "证书编号": st.column_config.TextColumn("证书编号", width="small"),
            "姓名": st.column_config.TextColumn("姓名", width="small"),
            "身份证号": st.column_config.TextColumn("身份证号", width="medium"),
            "培训日期": st.column_config.TextColumn("培训日期", width="medium"),
            "标准号": st.column_config.TextColumn("标准号", width="large"),
        }
    )
    temp_df = edited_df.drop(columns=["序号"])
    raw_data = temp_df.dropna(how='all').to_dict('records')
    data_to_process = [{k: str(v).strip() for k, v in row.items() if v is not None} for row in raw_data if any(row.values())]
else:
    c1, c2 = st.columns([2, 3])
    with c1:
        example_data = {"证书编号":["T-2025-001"],"姓名":["张三"],"身份证号":["440683..."],"培训日期":["2025年9月"],"标准号":["ISO9001"]}
        df_ex = pd.DataFrame(example_data)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_ex.to_excel(writer, index=False)
            ws = writer.sheets['Sheet1']
            for i in range(1, 6): ws.column_dimensions[get_column_letter(i)].width = 20
            for cell in ws[2]: cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        st.download_button("📥 下载标准模板", data=buf.getvalue(), file_name="学员信息上传模板.xlsx")
    with c2:
        up = st.file_uploader("上传文件", type=["xlsx", "csv"], label_visibility="collapsed")
        if up:
            df = pd.read_csv(up, dtype=str).fillna("") if up.name.endswith('.csv') else pd.read_excel(up, dtype=str).fillna("")
            data_to_process = [row for row in df.to_dict('records') if "示例" not in str(row.get('姓名',''))]
            if data_to_process: st.success(f"✅ 已加载 {len(data_to_process)} 条有效数据")

st.markdown("---")
st.markdown("### 第三步：模板确认与生成")
if os.path.exists(DEFAULT_TEMPLATE):
    t_opt = st.radio("模板选择：", ["使用内置模板", "上传本地新模板"], horizontal=True)
    t_path = DEFAULT_TEMPLATE if t_opt == "使用内置模板" else st.file_uploader("上传自定义模板", type=["docx"])
else:
    t_path = st.file_uploader("请上传 Word 模板", type=["docx"])

if t_path and data_to_process:
    if st.button("🚀 开始批量制作合并文档", use_container_width=True):
        try:
            master, bar, count = None, st.progress(0), 0
            for i, row in enumerate(data_to_process):
                name_v = str(row.get('姓名', '')).replace('nan', '').strip()
                if not name_v: continue
                count += 1
                doc = DocxTemplate(t_path)
                doc.render({'number': str(row.get('证书编号','')), 'name': name_v, 'id_card': str(row.get('身份证号','')), 'date': str(row.get('培训日期','')), 'standards': str(row.get('标准号',''))})
                tmp = io.BytesIO(); doc.save(tmp); tmp.seek(0)
                cur = Document(tmp)
                if master is None:
                    master = cur
                    composer = Composer(master)
                else:
                    master.add_page_break(); composer.append(cur)
                bar.progress((i + 1) / len(data_to_process))
            if master:
                out = io.BytesIO(); master.save(out); out.seek(0)
                st.balloons()
                st.download_button(f"🎁 下载汇总文档({count}份)", out.getvalue(), "证书汇总.docx", use_container_width=True)
        except Exception as e: st.error(f"制作失败：{e}")

# 渲染底部
st.markdown("---")
st.markdown(footer_html, unsafe_allow_html=True)
